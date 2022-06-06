
import openpyxl


inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
print(product_list.max_row)
# for product_row in product_list.max_row: